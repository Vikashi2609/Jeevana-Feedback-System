"""
xlsx_report.py — detailed Excel workbook per teacher (v1)

Workbook layout:
  Summary sheet:
    - Teacher / Total Responses / Classes Handled / Overall Average
    - Classes handled table
    - Overall question-wise average table (all classes combined)
    - Native bar chart of overall average per question
  One sheet per class/subject the teacher handles:
    - Question distribution (counts, %, average) — % and average are formulas
    - Native 100%-stacked bar chart of the rating distribution
    - Student comments for that class

Entry point (mirrors create_teacher_pdf in pdf_report.py):
    from xlsx_report import create_teacher_workbook
    create_teacher_workbook(teacher_data)

Expects the same `teacher_data` shape build_teacher_data() in processor.py
returns:
    {
      "teacher": str,
      "responses": int,
      "rows": [ {"class":.., "section":.., "subject":.., "q1":1-5, ..., "q10":1-5, "comment": str}, ... ],
      "classes": { (class, section, subject): [row, ...], ... },
      "distribution": [ {"question": "Q1", "ratings": {5:n,4:n,3:n,2:n,1:n}}, ... ],   # ALL classes combined
      "comments": { (class, section, subject): [comment, ...], ... },
    }
"""
import os
import re
import subprocess
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.worksheet.worksheet import Worksheet

# ---- Brand palette (matches pdf_report.py / chart.py) -------------------
NAVY = "1B3A5C"
GOLD = "D9A441"
LIGHT_BG = "F4F6F8"
TEXT = "33404D"
MUTED = "6B7684"

RATING_COLORS = {
    5: "3E8E70",
    4: "8FB99B",
    3: "F0C55E",
    2: "E19A3C",
    1: "C1554A",
}

FONT_NAME = "Arial"

thin = Side(style="thin", color="E2E6EA")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


# --------------------------------------------------------------------------
# Small style helpers
# --------------------------------------------------------------------------
def _header_cell(ws, row, col, text, fill=NAVY, color="FFFFFF", size=10, bold=True):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name=FONT_NAME, size=size, bold=bold, color=color)
    c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    return c


def _body_cell(ws, row, col, value, bold=False, align="center", fill=None, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=FONT_NAME, size=10, bold=bold, color=TEXT)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = BORDER
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if number_format:
        c.number_format = number_format
    return c


def _title_cell(ws, row, col, text, size=14, color=NAVY):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name=FONT_NAME, size=size, bold=True, color=color)
    return c


def _label_value(ws, row, label, value, number_format=None):
    lc = ws.cell(row=row, column=1, value=label)
    lc.font = Font(name=FONT_NAME, size=10, bold=True, color=MUTED)
    vc = ws.cell(row=row, column=2, value=value)
    vc.font = Font(name=FONT_NAME, size=11, bold=True, color=NAVY)
    if number_format:
        vc.number_format = number_format
    return vc


def _sanitize_sheet_name(name, used_names):
    # Excel sheet name rules: max 31 chars, no : \ / ? * [ ]
    clean = re.sub(r"[:\\/?*\[\]]", "", name).strip() or "Sheet"
    clean = clean[:31]
    base = clean
    i = 2
    while clean in used_names:
        suffix = f" ({i})"
        clean = (base[: 31 - len(suffix)] + suffix)
        i += 1
    used_names.add(clean)
    return clean


def _rating_avg_formula(rating_cols, total_cell):
    """SUMPRODUCT-based weighted average of ratings 5..1 against their counts."""
    # rating_cols: dict {5: 'B2', 4: 'C2', 3: 'D2', 2: 'E2', 1: 'F2'}
    terms = "+".join(f"{r}*{cell}" for r, cell in rating_cols.items())
    return f"=IFERROR(({terms})/{total_cell},0)"


def _add_distribution_chart(ws, anchor_cell, title, cat_ref, series_refs_with_titles,
                             stacked=True, height=8, width=16):
    chart = BarChart()
    chart.type = "bar"
    chart.grouping = "percentStacked" if stacked else "clustered"
    chart.overlap = 100 if stacked else -10
    chart.title = title
    chart.y_axis.title = None
    chart.x_axis.title = None
    chart.height = height
    chart.width = width
    chart.gapWidth = 60

    chart.set_categories(cat_ref)
    for series_ref, name, color in series_refs_with_titles:
        chart.add_data(series_ref, titles_from_data=False)
        s = chart.series[-1]
        s.tx = None
        from openpyxl.chart.series import SeriesLabel
        s.tx = SeriesLabel(v=name)
        s.graphicalProperties.solidFill = color
        s.graphicalProperties.line.noFill = True

    chart.legend.position = "r"
    ws.add_chart(chart, anchor_cell)
    return chart


# --------------------------------------------------------------------------
# Summary sheet
# --------------------------------------------------------------------------
def _build_summary_sheet(wb, data, class_rows):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    for col, w in zip("ABCDEFGH", [22, 16, 16, 16, 16, 16, 16, 14]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:D1")
    _title_cell(ws, 1, 1, "JEEVANA SCHOOL", size=16)
    ws.merge_cells("A2:D2")
    _title_cell(ws, 2, 1, "Teacher Evaluation Report — Summary", size=11, color=MUTED)

    _label_value(ws, 4, "Teacher", data["teacher"])
    _label_value(ws, 5, "Subject(s)", data.get("subject_line", ""))
    _label_value(ws, 6, "Total Responses", data["total_responses"])
    _label_value(ws, 7, "Classes Handled", len(data["classes"]))
    # Overall average is filled as a formula once the question table below exists;
    # placeholder written now, patched after that table is built.
    overall_avg_cell = ws.cell(row=8, column=2)
    ws.cell(row=8, column=1, value="Overall Average").font = Font(
        name=FONT_NAME, size=10, bold=True, color=MUTED)

    # ---- Classes handled table ----
    start = 11
    ws.cell(row=start - 1, column=1, value="Classes Handled").font = Font(
        name=FONT_NAME, size=12, bold=True, color=NAVY)
    headers = ["Class", "Section", "Subject", "Responses"]
    for i, h in enumerate(headers):
        _header_cell(ws, start, 1 + i, h)
    for i, c in enumerate(class_rows):
        r = start + 1 + i
        fill = LIGHT_BG if i % 2 else "FFFFFF"
        _body_cell(ws, r, 1, c["class"], fill=fill)
        _body_cell(ws, r, 2, c["section"], fill=fill)
        _body_cell(ws, r, 3, c["subject"], fill=fill)
        _body_cell(ws, r, 4, c["responses"], fill=fill)

    # ---- Overall question ratings table (all classes combined) ----
    q_start = start + len(class_rows) + 3
    ws.cell(row=q_start - 1, column=1, value="Overall Question Ratings (All Classes)").font = Font(
        name=FONT_NAME, size=12, bold=True, color=NAVY)
    headers = ["Question", "5", "4", "3", "2", "1", "Total", "Average"]
    for i, h in enumerate(headers):
        _header_cell(ws, q_start, 1 + i, h)

    q_first_data_row = q_start + 1
    for i, item in enumerate(data["distribution"]):
        r = q_first_data_row + i
        ratings = item["ratings"]
        fill = LIGHT_BG if i % 2 else "FFFFFF"
        _body_cell(ws, r, 1, item["question"], bold=True, fill=fill)
        for j, k in enumerate([5, 4, 3, 2, 1]):
            _body_cell(ws, r, 2 + j, ratings.get(k, 0), fill=fill)
        total_cell = f"G{r}"
        ws.cell(row=r, column=7, value=f"=SUM(B{r}:F{r})")
        ws.cell(row=r, column=7).font = Font(name=FONT_NAME, size=10, color=TEXT)
        ws.cell(row=r, column=7).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=7).border = BORDER
        if fill != "FFFFFF":
            ws.cell(row=r, column=7).fill = PatternFill("solid", fgColor=fill)

        rating_cols = {5: f"B{r}", 4: f"C{r}", 3: f"D{r}", 2: f"E{r}", 1: f"F{r}"}
        avg_formula = _rating_avg_formula(rating_cols, total_cell)
        avg_cell = ws.cell(row=r, column=8, value=avg_formula)
        avg_cell.number_format = "0.0"
        avg_cell.font = Font(name=FONT_NAME, size=10, bold=True, color=NAVY)
        avg_cell.alignment = Alignment(horizontal="center")
        avg_cell.border = BORDER
        if fill != "FFFFFF":
            avg_cell.fill = PatternFill("solid", fgColor=fill)

    q_last_data_row = q_first_data_row + len(data["distribution"]) - 1

    # patch overall average (mean of the per-question averages) now that range exists
    overall_avg_cell.value = f"=IFERROR(AVERAGE(H{q_first_data_row}:H{q_last_data_row}),0)"
    overall_avg_cell.number_format = "0.0"

    # ---- Native chart: average rating per question ----
    if len(data["distribution"]) > 0:
        cats = Reference(ws, min_col=1, min_row=q_first_data_row, max_row=q_last_data_row)
        vals = Reference(ws, min_col=8, min_row=q_start, max_row=q_last_data_row)  # includes header for name
        chart = BarChart()
        chart.type = "bar"
        chart.title = "Overall Average Rating per Question"
        chart.y_axis.title = None
        chart.x_axis.title = "Average (out of 5)"
        chart.height = 10
        chart.width = 16
        chart.gapWidth = 50
        chart.add_data(vals, titles_from_data=True)
        chart.set_categories(cats)
        chart.series[0].graphicalProperties.solidFill = NAVY
        chart.series[0].graphicalProperties.line.noFill = True
        chart.legend = None
        chart.x_axis.scaling.min = 0
        chart.x_axis.scaling.max = 5
        anchor = f"J{start}"
        ws.add_chart(chart, anchor)

    return q_last_data_row


# --------------------------------------------------------------------------
# Per-class sheet
# --------------------------------------------------------------------------
def _build_class_sheet(wb, sheet_name, class_key, rows, comments):
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_view.showGridLines = False
    cls, section, subject = class_key

    for col, w in zip("ABCDEFGHIJKLM", [10, 8, 8, 8, 8, 8, 9, 9, 8, 8, 8, 8, 8]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:D1")
    _title_cell(ws, 1, 1, f"{cls}-{section} : {subject}", size=13)
    _label_value(ws, 2, "Responses", len(rows))

    # ---- per-question distribution, computed from this class's raw rows ----
    start = 4
    headers = ["Question", "5", "4", "3", "2", "1", "Total", "Average",
               "%5", "%4", "%3", "%2", "%1"]
    for i, h in enumerate(headers):
        _header_cell(ws, start, 1 + i, h, size=9)

    n_questions = 10
    first_data_row = start + 1
    for qi in range(1, n_questions + 1):
        r = first_data_row + qi - 1
        counts = {5: 0, 4: 0, 3: 0, 2: 0, 1: 0}
        for row in rows:
            counts[row[f"q{qi}"]] += 1

        fill = LIGHT_BG if qi % 2 == 0 else "FFFFFF"
        _body_cell(ws, r, 1, f"Q{qi}", bold=True, fill=fill)
        for j, k in enumerate([5, 4, 3, 2, 1]):
            _body_cell(ws, r, 2 + j, counts[k], fill=fill)

        total_cell = f"G{r}"
        ws.cell(row=r, column=7, value=f"=SUM(B{r}:F{r})")
        ws.cell(row=r, column=7).font = Font(name=FONT_NAME, size=9.5, color=TEXT)
        ws.cell(row=r, column=7).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=7).border = BORDER
        if fill != "FFFFFF":
            ws.cell(row=r, column=7).fill = PatternFill("solid", fgColor=fill)

        rating_cols = {5: f"B{r}", 4: f"C{r}", 3: f"D{r}", 2: f"E{r}", 1: f"F{r}"}
        avg_cell = ws.cell(row=r, column=8, value=_rating_avg_formula(rating_cols, total_cell))
        avg_cell.number_format = "0.0"
        avg_cell.font = Font(name=FONT_NAME, size=9.5, bold=True, color=NAVY)
        avg_cell.alignment = Alignment(horizontal="center")
        avg_cell.border = BORDER
        if fill != "FFFFFF":
            avg_cell.fill = PatternFill("solid", fgColor=fill)

        for j, k in enumerate([5, 4, 3, 2, 1]):
            pct_col = 9 + j  # I..M
            src_col = get_column_letter(2 + j)  # B..F
            formula = f"=IFERROR({src_col}{r}/{total_cell},0)"
            pc = ws.cell(row=r, column=pct_col, value=formula)
            pc.number_format = "0%"
            pc.font = Font(name=FONT_NAME, size=9.5, color=TEXT)
            pc.alignment = Alignment(horizontal="center")
            pc.border = BORDER
            if fill != "FFFFFF":
                pc.fill = PatternFill("solid", fgColor=fill)

    last_data_row = first_data_row + n_questions - 1

    # ---- 100%-stacked distribution chart, using the % columns ----
    cats = Reference(ws, min_col=1, min_row=first_data_row, max_row=last_data_row)
    chart = BarChart()
    chart.type = "bar"
    chart.grouping = "percentStacked"
    chart.overlap = 100
    chart.title = "Rating Distribution by Question"
    chart.height = 9
    chart.width = 17
    chart.gapWidth = 50
    chart.legend.position = "r"

    # add series in 1..5 order (bottom to top) so 5-star sits at the far end, matching PDF legend order
    for j, rating in enumerate([1, 2, 3, 4, 5]):
        col = 9 + (5 - rating)  # I=%5,J=%4,K=%3,L=%2,M=%1 -> map rating to its column
        ref = Reference(ws, min_col=col, min_row=start, max_row=last_data_row)  # include header row for name
        chart.add_data(ref, titles_from_data=True)
        s = chart.series[-1]
        s.graphicalProperties.solidFill = RATING_COLORS[rating]
        s.graphicalProperties.line.noFill = True
    chart.set_categories(cats)

    ws.add_chart(chart, f"A{last_data_row + 3}")

    # ---- comments ----
    comments_start = last_data_row + 3 + 20  # leave room below the chart
    ws.cell(row=comments_start, column=1, value="Student Comments").font = Font(
        name=FONT_NAME, size=12, bold=True, color=NAVY)
    if comments:
        for i, comment in enumerate(comments):
            r = comments_start + 1 + i
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            c = ws.cell(row=r, column=1, value=f"•  {comment}")
            c.font = Font(name=FONT_NAME, size=10, color=TEXT)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    else:
        r = comments_start + 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        c = ws.cell(row=r, column=1, value="No written comments submitted.")
        c.font = Font(name=FONT_NAME, size=10, italic=True, color=MUTED)


# --------------------------------------------------------------------------
# Entry point
# --------------------------------------------------------------------------
def create_teacher_workbook(teacher_data, out_dir="exports/Excel", recalc_script_path=None):
    """
    Entry point to call from main.py:
        from xlsx_report import create_teacher_workbook
        create_teacher_workbook(teacher_data)

    Note: formulas are written but not pre-calculated here — Excel recalculates
    them automatically the moment the file is opened, so this is only relevant
    if you need cached values *without* opening the file in Excel first (e.g.
    reading it back with pandas). Pass recalc_script_path to a LibreOffice-based
    recalculation script if you need that.
    """
    teacher = teacher_data["teacher"]
    raw_classes = teacher_data.get("classes") or {}
    raw_comments = teacher_data.get("comments") or {}

    class_rows = []
    for (cls, section, subject), rows in raw_classes.items():
        class_rows.append({"class": cls, "section": section, "subject": subject,
                            "responses": len(rows), "_key": (cls, section, subject), "_rows": rows})
    class_rows.sort(key=lambda c: (str(c["class"]), str(c["section"]), str(c["subject"])))

    subjects = []
    for c in class_rows:
        if c["subject"] not in subjects:
            subjects.append(c["subject"])

    data = {
        "teacher": teacher,
        "subject_line": ", ".join(subjects),
        "total_responses": teacher_data.get("responses", sum(c["responses"] for c in class_rows)),
        "classes": class_rows,
        "distribution": teacher_data.get("distribution") or [],
    }

    wb = Workbook()
    _build_summary_sheet(wb, data, class_rows)

    used_names = {"Summary"}
    for c in class_rows:
        sheet_name = _sanitize_sheet_name(f"{c['class']}-{c['section']} {c['subject']}", used_names)
        comments = raw_comments.get(c["_key"], [])
        _build_class_sheet(wb, sheet_name, c["_key"], c["_rows"], comments)

    os.makedirs(out_dir, exist_ok=True)
    safe_name = "".join(ch for ch in teacher if ch.isalnum() or ch in (" ", "_", "-")).strip()
    out_path = os.path.join(out_dir, f"{safe_name}.xlsx")
    wb.save(out_path)

    if recalc_script_path and os.path.exists(recalc_script_path):
        subprocess.run([sys.executable, recalc_script_path, out_path], check=False)

    return out_path